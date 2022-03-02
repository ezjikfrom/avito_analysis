#Work with Avito
from concurrent.futures.thread import _worker
import openpyxl
import requests
from bs4 import BeautifulSoup as BS
from time import sleep
import random
#load to Google
import pandas as pd # внешняя библиотека. Нужна для удобного работа с данными
import gspread #работа с гугл таблицами
from gspread_dataframe import set_with_dataframe #работа с гугл таблицами
#work with dates
from fake_useragent import UserAgent


class AvitoStatistic():
    #Поиск платных и бесплатных объявлений
    def findX(self, html, id):
        x = html.split(id)
        for i in range(html.count(id)+1):
            bigText = x[i][:410]
            if 'https%3A%2F%2Fwww.avito.st%2Fs%2Fcommon%2Fcomponents%2Fmonetization%2Ficons%2Fweb%2F' in bigText:
                index = bigText.find('https%3A%2F%2Fwww.avito.st%2Fs%2Fcommon%2Fcomponents%2Fmonetization%2Ficons%2Fweb%2F')
                xt = bigText[index+84:].split('svg')[0]+'svg'
                return xt
                break
 
    def cityCovert(self, city, google_sheet_url,table = "Города"):
        #wb = openpyxl.load_workbook('citys.xlsx')
        #ws = wb.active
        #for i in range(1, ws.max_row+1):
        #    if city.lower() == ws['A'+str(i)].value.lower():
        #        return ws['B'+str(i)].value.lower()
        #        break
        # ACCES GOOGLE SHEET
        num = random.choice(['3','4'])
        print('1-',num)
        gc = gspread.service_account(filename='credentials{}.json'.format(num))
        
        sh = gc.open_by_url(google_sheet_url)
        
        # table - это какую таблицу в гугл документе открываем. Может быть порядковый номер ее, а может просто название
        if type(table) == int:
            worksheet = sh.get_worksheet(table) #-> 0 - first sheet, 1 - second sheet etc. 
        else:
            worksheet = sh.worksheet(table)  
        data = worksheet.get_all_values()
        #print(city)
        for i in range(0, len(data)):
            if data[i][0].lower() == city.lower():   
                return str(data[i][1])
    #Ссылки на страниы объявлений
    def get_links(self, html):
        cleen = []
        x2 = []
        x5 = []
        x10 = []
        all = []
        soup = BS(html.text, 'html.parser')
        bloks = soup.find_all('div', class_='iva-item-root-_lk9K')
        for i in bloks:
            link = 'https://www.avito.ru'+i.find('a', class_='link-link-MbQDP').get('href')
            blok = i.find('div', class_='iva-item-dateInfoStep-_acjp')
            if blok.find('div', class_='styles-arrow-jfRdd'):
                #print(html.text)
                x = self.findX(html.text, link.split('_')[-1])
                if 'x2' in x:
                    x2.append(link)
                elif 'x5' in x:
                    x5.append(link)
                elif 'x10' in x:
                    x10.append(link)
            else:
                cleen.append(link)
        all.append(x10)
        all.append(x5)
        all.append(x2)
        all.append(cleen)
        return all
    #################
    def worker(self, zapros, city,categoria,google_sheet_url):
        HOST = 'https://www.avito.ru/'
        
        headers = {'user-agent': str(UserAgent().random)}
        ###############
        #input
        page = 1
        city_eng = (self.cityCovert(city,google_sheet_url)).lower()
        ###############
        #LINK = HOST + city_eng + '?q=' + zapros
        
        metki = [[],[],[],[]]
        client = requests.Session()
        for i in range(1,2):
            if categoria != '':
                if i == 1:
                    LINK = HOST + city_eng + '/'  + categoria + '?q=' + zapros
                else:
                    LINK = HOST + city_eng + '/'  + categoria + '?p=' + str(i) + '&q=' + zapros
            else:
                if i == 1:
                    LINK = HOST + city_eng + '?q=' + zapros
                else:
                    LINK = HOST + city_eng + '?p=' + str(i) + '&q=' + zapros
            print(LINK)
            html = client.get(LINK, headers=headers)
            print(html)
            #print(html.text)
            #for key, value in html.request.headers.items():
                #print(key+": "+value)
            m = self.get_links(html)
            for n in range(4):
                metki[n].extend(m[n])

            sleep(random.uniform(4.2, 7.1))
            soup = BS(html.text, 'html.parser')
            if i == 1:
                try:
                    allAds = soup.find('span', class_='page-title-count-wQ7pG').get_text(strip=True)
                except:
                    allAds = '0'

            if not soup.find('div', class_='js-pages'):
                #print('Только', i ,'страниц')
                break
        ####################
        dataDay = [[],[],[],[]]
        dataAll = [[],[],[],[]]
        ###################
        #Собираем просмотры
        counter = len(metki[0])+len(metki[1])+len(metki[2])+len(metki[3])
        for i in range(4):
            if len(metki[i]) != 0:
                day = 0
                all = 0
                client = requests.Session()
                for n in range(len(metki[i])):
                    #print('****************')
                    counter-=1
                    #print('Осталось', counter)
                    
                    if city_eng in metki[i][n]:
                        print(metki[i][n])
                        html = client.get(metki[i][n], headers=headers)
                        print('2:\n',html)
                        soup = BS(html.text, 'html.parser')
                        try:
                            view = soup.find('div', class_='title-info-metadata-item title-info-metadata-views').get_text(strip=True)
                            all = all + int(view.split(' ')[0])
                            if len(view.split(' ')[1])!=0:
                                day = day + int(view.split(' ')[1].replace('(','').replace('+','').replace(')',''))
                            #else:
                                #None
                            sleep(random.uniform(4.4, 7.6))
                        except Exception as mist:
                            print('mist in 153', mist)
                            #day = 0
                            #all = 0
                    else:
                        day = 0
                        all = 0
                dataDay[i] = day/len(metki[i])
                dataAll[i] = all/len(metki[i])
            else:
                dataDay[i] = 0
                dataAll[i] = 0
        ####################

        
        bigData = ({
                'Запрос': zapros,
                'Город': city,
                'Всего объявлений': allAds.replace(u'\xa0',''),
                'Всего х10': len(metki[0]),
                'Всего х5': len(metki[1]),
                'Всего х2': len(metki[2]),
                'Всего бесплатных': len(metki[3]),
                'В среднем просмотров за сегодня х10': dataDay[0],
                'В среднем просмотров за сегодня х5': dataDay[1],
                'В среднем просмотров за сегодня х2': dataDay[2],
                'В среднем просмотров за сегодня бесплатных': dataDay[3],
                'В среднем просмотров за всё время х10': dataAll[0],
                'В среднем просмотров за всё время х5': dataAll[1],
                'В среднем просмотров за всё время х2': dataAll[2],
                'В среднем просмотров за всё время бесплатных': dataAll[3],
                })
        return bigData
    def init(self):
        pass
###########
class LoadToGoogle():
    #выгрузка за раз целого блока благодаря pandas (тебе так и нужно будет записывать)
    def load_to_google(self,data,google_sheet_url,table='Результат анализа'):
        # ACCES GOOGLE SHEET
        gc = gspread.service_account(filename='credentials.json')
        #sh = gc.open_by_key('your_google_sheet_ID')
        sh = gc.open_by_url(google_sheet_url)
        # table - это какую таблицу в гугл документе открываем. Может быть порядковый номер ее, а может просто название
        if type(table) == int:
            worksheet = sh.get_worksheet(table) #-> 0 - first sheet, 1 - second sheet etc. 
        else:
            worksheet = sh.worksheet(table)  
        
        worksheet.clear()# это очищает всю таблицу целиком  
        # APPEND DATA TO SHEET - добавляем целиком за раз благоадря формату от pandas (библиотеку подключали выше)

        set_with_dataframe(worksheet, data) #-> THIS EXPORTS YOUR DATAFRAME TO THE GOOGLE SHEET

    def load_to_google_one_string(self,data,google_sheet_url,table='Вводная инфа'):
        # ACCES GOOGLE SHEET
        gc = gspread.service_account(filename='credentials.json')
        #sh = gc.open_by_key('your_google_sheet_ID')
        sh = gc.open_by_url(google_sheet_url)
        # table - это какую таблицу в гугл документе открываем. Может быть порядковый номер ее, а может просто название
        if type(table) == int:
            worksheet = sh.get_worksheet(table) #-> 0 - first sheet, 1 - second sheet etc. 
        else:
            worksheet = sh.worksheet(table)  
        
        worksheet.update('D2', data)

    def init(self):
        pass
###########
class Load_from_Google():
    def load_from_google(self,google_sheet_url,table='Вводная инфа'):
        # ACCES GOOGLE SHEET
        num = random.choice(['3','4'])
        print('2-',num)
        gc = gspread.service_account(filename='credentials{}.json'.format(num))
        #sh = gc.open_by_key('your_google_sheet_ID')
        sh = gc.open_by_url(google_sheet_url)
        # table - это какую таблицу в гугл документе открываем. Может быть порядковый номер ее, а может просто название
        if type(table) == int:
            worksheet = sh.get_worksheet(table) #-> 0 - first sheet, 1 - second sheet etc. 
        else:
            worksheet = sh.worksheet(table)  

        data = worksheet.get_all_values()
        return data
    def init(self):
        pass
###########
#Ввод данных
mist_time = 10
while True:
    
    #try:
    google_sheet_url = 'https://docs.google.com/spreadsheets/d/14tjX7k_4HxatWWekaw3ToRHHCqUduCMb3R7k4fhJa6s/edit#gid=0'
    ld = Load_from_Google()
    basa = ld.load_from_google(google_sheet_url)
    #print(basa)
    if basa[1][3] == 'Нужно' or basa[1][3] == 'Идет анализ (не трогать)':
        writing = LoadToGoogle()
        # write 'Идет анализ (не трогать)'
        data = 'Идет анализ (не трогать)'
        writing.load_to_google_one_string(data,google_sheet_url)

        ran = AvitoStatistic()
        
        data = ({
                        'Запрос': [],
                        'Город': [],
                        'Всего объявлений': [],
                        'Всего х10': [],
                        'Всего х5': [],
                        'Всего х2': [],
                        'Всего бесплатных': [],
                        'В среднем просмотров за сегодня х10': [],
                        'В среднем просмотров за сегодня х5': [],
                        'В среднем просмотров за сегодня х2': [],
                        'В среднем просмотров за сегодня бесплатных': [],
                        'В среднем просмотров за всё время х10': [],
                        'В среднем просмотров за всё время х5': [],
                        'В среднем просмотров за всё время х2': [],
                        'В среднем просмотров за всё время бесплатных': [],
                        })
        data = pd.DataFrame(data = data)
        #print(data)
        #print(basa)
        sort_basa = []
        for num in range(len(basa)):
            if basa[num][0] != '' and basa[num][1] !='':
                sort_basa.append(basa[num])
        #print(basa)
        #print(sort_basa)
        for i in range(1,len(sort_basa)):
            #print(basa[i][0], basa[i][1])
            newdata = ran.worker(basa[i][0], basa[i][1], basa[i][2],google_sheet_url)
            newdata = pd.DataFrame(data = newdata, index = [0])
            data = pd.concat([data,newdata], ignore_index=True)
            writing.load_to_google(data, google_sheet_url)
        #result = writing.load_to_google(data, google_sheet_url)
        # пишем в гугл таблицу, что проанализировали
        data = 'Проанализировано'
        writing.load_to_google_one_string(data,google_sheet_url)
    elif basa[1][3] == 'Проанализировано':
        sleep(60)
    '''except Exception as mist:
        mist_time += 1
        print(mist)
        print('Mistake, sleep',mist_time)
        
        sleep(mist_time)'''

